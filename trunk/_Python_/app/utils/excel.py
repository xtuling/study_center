# -*-coding:utf8-*-

"""
@author：three
@note: Excel工具
"""
import xlsxwriter
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook

import xlrd
import xlwt


class MyExcelReader:
    """excel 读取"""
    sheet = WorkBook = None
    rows = 0
    cols = 0

    def row_number(self):
        """获得行总数"""
        return self.rows

    def col_number(self):
        """获得列总数"""
        return self.cols


class MyExcelWriter:
    """excel 写入"""
    WorkBook = None
    path = None

    def save(self):
        """save to file"""
        self.WorkBook.save(self.path)

    def create_sheet(self, sheet):
        return None


class Excel2003Reader(MyExcelReader):
    """ 读取2003格式 """

    def __init__(self, path, sheet=0):
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(sheet)

        self.sheet = sh
        self.WorkBook = wb
        self.rows = self.sheet.nrows
        self.cols = self.sheet.ncols

    def read(self, start=0, end=0):
        data = []
        if start > end or start >= self.rows:
            return data

        if 0 == end or end > self.rows:
            end = self.rows

        for index in range(start, end):
            data.append(self.sheet.row_values(index))
        return data

    def read_row(self, num):
        if num > self.rows:
            return None
        return self.sheet.row_values(num)

    def read_column(self, num):
        if num > self.cols:
            return None
        return self.sheet.col_values(num)

    def read_cell(self, x, y):
        """读取单元格"""
        return self.sheet.cell(x, y).value

    def set_sheet(self, sheet):
        """重新选择sheet"""
        sh = self.WorkBook.sheet_by_index(sheet)

        self.sheet = sh
        self.rows = self.sheet.nrows
        self.cols = self.sheet.ncols


class Excel2003Writer(MyExcelWriter):
    """Excel2003 格式写入"""

    def __init__(self, path):
        """__init__ documentation"""
        self.path = path
        self.WorkBook = xlwt.Workbook()

    def create_font(self, attr):
        """创建字体属性"""
        fonts = []
        for k in attr:
            if 'bold' == k:
                fonts.append('bold ' + attr[k])
            elif 'underline' == k:
                fonts.append('underline ' + attr[k])
            elif 'italic' == k:
                fonts.append('italic ' + attr[k])
            elif 'height' == k:
                fonts.append('height ' + attr[k])
            elif 'width' == k:
                fonts.append('width ' + attr[k])
        font = ''
        if 0 < len(fonts):
            font = 'font: ' + ', '.join(fonts) + ';'
        return font

    def create_pattern(self, attr):
        patterns = []
        for k in attr:
            if 'pattern' == k:
                patterns.append('pattern ' + attr[k])
            elif 'fore_colour' == k:
                patterns.append('fore_colour ' + attr[k])
        pattern = ''
        if 0 < len(patterns):
            pattern = 'pattern: ' + ', '.join(patterns) + ';'
        return pattern

    def create_align(self, attr):
        aligns = []
        for k in attr:
            if 'wrap' == k:
                aligns.append('wrap ' + attr[k])
        align = ''
        if 0 < len(aligns):
            align = 'align: ' + ', '.join(aligns) + ';'
        return align

    def create_sheet(self, sheet):
        """创建工作薄"""
        return self.WorkBook.add_sheet(sheet)

    def set_cell(self, sheet, x, y, value):
        """设置表格"""
        if isinstance(value, dict):
            val = value['value']
            styles = []
            for k in value:
                if 'font' == k:
                    styles.append(self.create_font(value[k]))
                elif 'pattern' == k:
                    styles.append(self.create_pattern(value[k]))
                elif 'align' == k:
                    styles.append(self.create_align(value[k]));
            if 'write_merge' in value.keys():
                height = 0
                if 'height' in value['write_merge'].keys():
                    height = value['write_merge']['height']
                width = 0
                if 'width' in value['write_merge'].keys():
                    width = value['write_merge']['width']
                print('height:', height)
                print('width:', width)
                sheet.write_merge(x, x + height, y, y + width, val, xlwt.Style.easyxf(''.join(styles)))
            else:
                sheet.write(x, y, val, xlwt.Style.easyxf(''.join(styles)))
        else:
            sheet.write(x, y, value)


class Excel2007Reader(MyExcelReader):
    """Excel2007 格式读取"""

    def __init__(self, path, sheet):
        """init """
        wb = load_workbook(filename=path, read_only=True)
        sheet_name = wb.get_sheet_names()[sheet]
        self.sheet = wb.get_sheet_name(name=sheet_name)
        self.cols = self.sheet.get_highest_column()
        self.rows = self.sheet.get_highest_row()

    def read(self, start=0, end=0):
        data = []
        if start > end or start >= self.rows:
            return data

        if 0 == end or end > self.rows:
            end = self.rows

        for index in range(start, end):
            data.append(self.sheet.rows(index))
        return data

    def read_cell(self, x, y):
        """cell"""
        return self.sheet.cell(rows=x, column=y).value

    def read_row(self, x):
        """row"""
        return self.sheet.rows[x]

    def read_column(self, y):
        return self.sheet.cols[y]


class Excel2007Writer(MyExcelWriter):
    def __init__(self, path):
        """"""
        self.path = path
        # 使用openpyxl写入
        self.WorkBook = Workbook(write_only=True)

    def create_sheet(self, sheet):
        return self.WorkBook.create_sheet(sheet)

    def set_cell(self, sheet, x, y, value):
        sheet.cell(x, y, value)

    def append_row(self, sheet, row):
        sheet.append(row)


class Excel2007Writer2(MyExcelWriter):
    """excel 2007 写入，使用 xlsxWriter"""

    def __init__(self, path):
        """"""
        self.path = path
        # Create an new Excel file and add a worksheet.
        self.WorkBook = xlsxwriter.Workbook(path)

    def create_sheet(self, sheet):
        return self.WorkBook.add_worksheet(sheet)

    def set_cell(self, sheet, x, y, value):
        sheet.write(x, y, value)

    def sheet_format(self, sheet, style):
        sheet.add_format(style)

    def insert_image(self, sheet, path, style):
        sheet.insert_image(style, path)

    def save(self):
        self.WorkBook.close()
